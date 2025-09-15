# Đường dẫn: excel_toolkit/utils/data_ops.py
# Phiên bản 2.1 - Khôi phục đầy đủ tính năng cho hàm df_read
# Ngày cập nhật: 2025-09-12

import logging
import pandas as pd
import openpyxl as opx
import xlwings as xw
import os
import csv

# ======================================================================
# --- Nhóm 1: Đọc dữ liệu cấp cao bằng Pandas (Đầy đủ tính năng) ---
# ======================================================================

def df_read(data_input, sheet_name=0, header_row=0, use_cols=None, flt=None, output=None, chunksize=None):
    """
    Đọc dữ liệu từ nhiều định dạng file khác nhau và trả về dưới dạng DataFrame.
    Hàm này hỗ trợ đọc từ file Excel, CSV, và cả DataFrame đã có.
    Nó cũng cho phép tùy chỉnh các tham số như hàng tiêu đề, sheet, cột, 
    bộ lọc, định dạng đầu ra và đọc file theo chunk.
    """
    logging.debug(f"Bắt đầu quy trình df_read với data_input: {data_input}.")

    # --- Hàm nội bộ để kiểm tra và đọc dữ liệu từ nguồn ---
    def _read_source_data(source, sheet, header, cols, chunk):
        if isinstance(source, pd.DataFrame):
            return source[cols] if cols else source

        if not os.path.isfile(source):
            raise FileNotFoundError(f"File không tồn tại: {source}.")

        file_extension = os.path.splitext(source)[1].lower()
        
        try:
            if file_extension in ['.xlsx', '.xlsm', '.xls']:
                return pd.read_excel(source, sheet_name=sheet, header=header, usecols=cols)
            elif file_extension == '.csv':
                return pd.read_csv(source, sep=',', header=header, usecols=cols, on_bad_lines='skip', quoting=csv.QUOTE_MINIMAL, chunksize=chunk)
            else:
                raise ValueError(f"Định dạng file không được hỗ trợ: {file_extension}.")
        except Exception as e:
            logging.error(f"Lỗi khi đọc file '{source}': {e}")
            raise

    # --- Hàm nội bộ để áp dụng bộ lọc ---
    def _apply_filter(df, filters):
        if not filters:
            return df
        
        logging.debug(f"Bắt đầu áp dụng bộ lọc: {filters}")
        filtered_df = df.copy()
        for col, value, op in filters:
            if col not in filtered_df.columns:
                raise KeyError(f"Cột '{col}' không tồn tại trong DataFrame.")
            
            if op == 'exact':
                filtered_df = filtered_df[filtered_df[col] == value]
            elif op == 'contains':
                filtered_df = filtered_df[filtered_df[col].astype(str).str.contains(str(value), na=False)]
            elif op == '>':
                filtered_df = filtered_df[pd.to_numeric(filtered_df[col], errors='coerce') > value]
            elif op == '<':
                filtered_df = filtered_df[pd.to_numeric(filtered_df[col], errors='coerce') < value]
            elif op == '>=':
                filtered_df = filtered_df[pd.to_numeric(filtered_df[col], errors='coerce') >= value]
            elif op == '<=':
                filtered_df = filtered_df[pd.to_numeric(filtered_df[col], errors='coerce') <= value]
            elif op == '!=':
                filtered_df = filtered_df[filtered_df[col] != value]
            else:
                raise ValueError(f"Toán tử không hợp lệ: {op}.")
        
        logging.info(f"DataFrame sau khi lọc có shape: {filtered_df.shape}")
        return filtered_df

    # --- Hàm nội bộ để xử lý đầu ra ---
    def _handle_output(df, out_format):
        if not out_format:
            return df

        logging.debug(f"Bắt đầu xử lý output với định dạng: {out_format}")
        if out_format == "list":
            return df.values.tolist()
        elif out_format == "dict":
            return df.to_dict(orient='records')
        
        # Nếu out_format là một đường dẫn file
        try:
            output_ext = os.path.splitext(out_format)[1].lower()
            if output_ext == '.csv':
                df.to_csv(out_format, index=False)
            elif output_ext in ['.xlsx', '.xls']:
                df.to_excel(out_format, index=False)
            else:
                raise ValueError(f"Định dạng file output không được hỗ trợ: {output_ext}.")
            logging.info(f"Đã ghi DataFrame thành công ra file: {out_format}")
            return df # Trả về df để có thể tiếp tục dùng
        except Exception as e:
            logging.error(f"Lỗi khi ghi output ra file '{out_format}': {e}")
            raise

    # --- Luồng thực thi chính ---
    try:
        df = _read_source_data(data_input, sheet_name, header_row, use_cols, chunksize)
        df_filtered = _apply_filter(df, flt)
        result = _handle_output(df_filtered, output)
        logging.info("Hoàn tất quy trình df_read thành công.")
        return result
    except Exception as e:
        logging.error(f"Lỗi trong quy trình df_read: {e}")
        return None

# ======================================================================
# --- Nhóm 2: Đọc dữ liệu bằng các engine khác ---
# ======================================================================

def read_with_openpyxl(file_path, sheet_name, read_only=True):
    """
    Đọc toàn bộ dữ liệu từ một sheet bằng openpyxl.
    Phương pháp này rất nhanh và không cần mở ứng dụng Excel,
    lý tưởng cho việc trích xuất dữ liệu thô.
    """
    logging.debug(f"Bắt đầu đọc dữ liệu bằng openpyxl từ file '{file_path}', sheet '{sheet_name}'.")
    try:
        if not os.path.exists(file_path):
            logging.error(f"Lỗi: File không tồn tại tại đường dẫn '{file_path}'.")
            return None

        workbook = opx.load_workbook(filename=file_path, read_only=read_only)
        if sheet_name not in workbook.sheetnames:
            logging.error(f"Lỗi: Không tìm thấy sheet '{sheet_name}' trong file.")
            workbook.close()
            return None
            
        sheet = workbook[sheet_name]
        data = [[cell.value for cell in row] for row in sheet.iter_rows()]
        
        workbook.close()
        logging.info(f"Đã đọc thành công {len(data)} hàng dữ liệu bằng openpyxl.")
        return data
    except Exception as e:
        logging.error(f"Lỗi khi đọc file bằng openpyxl: {e}")
        return None

def read_with_xlwings(wb, sheet_name, as_df=False):
    """
    Đọc dữ liệu từ một workbook object của xlwings đã được mở sẵn.
    """
    logging.debug(f"Bắt đầu đọc dữ liệu bằng xlwings từ sheet '{sheet_name}'.")
    try:
        sheet = wb.sheets[sheet_name]
        if as_df:
            data = sheet.used_range.options(pd.DataFrame, index=False, header=True).value
            logging.info(f"Đã đọc thành công dữ liệu vào DataFrame, shape: {data.shape}.")
        else:
            data = sheet.used_range.value
            logging.info(f"Đã đọc thành công dữ liệu dưới dạng list of lists.")
        return data
    except KeyError:
        logging.error(f"Lỗi: Không tìm thấy sheet '{sheet_name}'.")
        return None
    except Exception as e:
        logging.error(f"Lỗi khi đọc dữ liệu bằng xlwings: {e}")
        return None

