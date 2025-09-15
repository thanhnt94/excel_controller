# Đường dẫn: excel_toolkit/processes/delete_hidden_sheets.py
# Phiên bản 4.0 - Cập nhật để hoạt động với ExcelController đã mở sẵn
# Ngày cập nhật: 2025-09-15

import logging
import os
import openpyxl
from excel_controller import ExcelController

def _find_dependencies(file_path, visible_sheets, hidden_sheets):
    """
    Hàm nội bộ: Sử dụng openpyxl để tìm các công thức phụ thuộc vào sheet ẩn.
    """
    logging.info("Bắt đầu tìm kiếm các công thức phụ thuộc vào sheet ẩn...")
    dependencies = {}
    try:
        opx_wb = openpyxl.load_workbook(file_path, data_only=False)
        
        for sheet_name in visible_sheets:
            if sheet_name not in opx_wb.sheetnames:
                continue
            
            ws = opx_wb[sheet_name]
            dependent_cells = []
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' and cell.value:
                        formula = str(cell.value)
                        for hidden_sheet_name in hidden_sheets:
                            if f"'{hidden_sheet_name}'!" in formula or f"{hidden_sheet_name}!" in formula:
                                dependent_cells.append(cell.coordinate)
                                logging.debug(f"  -> Tìm thấy tham chiếu đến '{hidden_sheet_name}' trong ô {cell.coordinate} của sheet '{sheet_name}'")
                                break
            
            if dependent_cells:
                dependencies[sheet_name] = dependent_cells
                
    except Exception as e:
        logging.error(f"Lỗi khi tìm kiếm công thức phụ thuộc bằng openpyxl: {e}")
    
    if dependencies:
        logging.info(f"Đã tìm thấy {sum(len(v) for v in dependencies.values())} ô có công thức phụ thuộc.")
    else:
        logging.info("Không tìm thấy công thức nào phụ thuộc vào sheet ẩn.")
        
    return dependencies

def run(controller, file_path):
    """
    Xóa các sheet ẩn trong workbook một cách thông minh và an toàn.
    """
    logging.info(f"Bắt đầu quy trình xóa sheet ẩn (an toàn) cho file: {os.path.basename(file_path)}")
    try:
        # Lấy danh sách sheet trực tiếp từ controller
        visible_sheets, hidden_sheets = controller.get_sheets_visibility()
        
        if not hidden_sheets:
            logging.info("Không có sheet ẩn nào để xóa. Kết thúc quy trình.")
            return

        # Bước 1: Tìm các ô có công thức phụ thuộc
        dependencies = _find_dependencies(file_path, visible_sheets, hidden_sheets)
        
        # Bước 2: Thay thế công thức bằng giá trị, sử dụng controller
        if dependencies:
            logging.info("Bắt đầu thay thế công thức bằng giá trị...")
            for sheet_name, cell_addresses in dependencies.items():
                for address in cell_addresses:
                    try:
                        # Lấy giá trị và ghi lại bằng các phương thức của controller
                        current_value = controller.get_cell_value(sheet_name, address)
                        controller.set_cell_value(sheet_name, address, current_value)
                        logging.info(f"  -> Đã thay thế công thức trong ô {address} của sheet '{sheet_name}' bằng giá trị.")
                    except Exception as e:
                        logging.error(f"Lỗi khi thay thế công thức tại ô {address}: {e}")
            logging.info("Hoàn tất việc thay thế công thức.")

        # Bước 3: Xóa các sheet ẩn sau khi đã xử lý an toàn
        logging.info("Bắt đầu xóa các sheet ẩn...")
        for sheet_name in hidden_sheets:
            controller.delete_sheet(sheet_name)
        
        logging.info(f"Hoàn tất quy trình xóa sheet ẩn cho file: {os.path.basename(file_path)}")
        
    except Exception as e:
        logging.error(f"Lỗi trong quy trình xóa sheet ẩn cho file '{file_path}': {e}", exc_info=True)
        raise
